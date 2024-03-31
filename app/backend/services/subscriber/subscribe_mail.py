from flask import jsonify, request
from datetime import datetime
from sqlalchemy import and_

from backend.models.model import Events, insert_row, commit_session, rollback_session
from backend.models.database import db

def mail_subscription(user_data):
    try:
        import smtplib 
        from email.mime.multipart import MIMEMultipart 
        from email.mime.text import MIMEText 
        from email.mime.base import MIMEBase 
        from email import encoders 

        fromaddr = "ramscrush000@gmail.com"
        toaddr = user_data["email"]
        to_name = user_data["username"]
        event_type = user_data["domain"]

        from openpyxl import Workbook

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Define headers
        headers = ["Event Date", "Event Description", "Domain"]

        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Query events and write details to subsequent rows
        get_events = Events.query.filter(Events.event_type == event_type).order_by(Events.event_date).all()
        for row, event in enumerate(get_events, start=2):
            ws.cell(row=row, column=1, value=event.event_date)
            ws.cell(row=row, column=2, value=event.event_description)
            ws.cell(row=row, column=3, value=event.domain)

        # Save the workbook
        wb.save("event_details.xlsx")

        # instance of MIMEMultipart 
        msg = MIMEMultipart() 

        msg['From'] = fromaddr 

        msg['To'] = toaddr 

        msg['Subject'] = f"{event_type.title()} Events List"

        body = f"Hello {to_name}, PFA Below!"

        msg.attach(MIMEText(body, 'plain')) 

        filename = "event_details.xlsx"
        attachment = open(f"{filename}", "rb") 

        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {filename}")
        msg.attach(part)

        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(fromaddr, "hrcv ubko nhoi ibmx")
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            print("Sent mail")
 
        response = jsonify({'status': True, 'message': True})
        return response
    except Exception as e:
        rollback_session()
        return jsonify({"status": False, "error": str(e)})